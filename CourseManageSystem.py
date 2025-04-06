import csv
from typing import List, Tuple, Dict
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, column_index_from_string
import re
import logging
import os
from utils.db_manager import DatabaseManager

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class CourseManagementSystem:
    def __init__(self, db_name: str):
        self.db_name = db_name
        self.db_manager = DatabaseManager(db_name)
        self.db_manager.initialize()

    def _extract_sheet_number(self, sheet_name: str) -> int:
        """Extract sheet number from sheet name (e.g., 'd26' or 'Danh mục d26' -> 26)"""
        match = re.search(r'd(\d+)', sheet_name)
        return int(match.group(1)) if match else -1

    def _is_danh_muc_sheet(self, sheet_name: str) -> bool:
        """Check if sheet is a 'Danh mục' sheet"""
        return sheet_name.strip().startswith('Danh')

    def _is_d_sheet(self, sheet_name: str) -> bool:
        """Check if sheet is a 'd' sheet"""
        return sheet_name.strip().startswith('d')

    def _parse_course_details_sheet(self, sheet, sheet_name: str) -> List[List]:
        """Parse a single Danh mục sheet for course details"""
        course_details = []
        pages = self._detect_pages(sheet)
        logger.info(f"Found {len(pages)} pages in course details sheet: {sheet_name}")

        for page_num, (start_col_idx, end_col_idx, start_row_idx, end_row_idx) in enumerate(pages, 1):
            logger.info(f"Processing page {page_num} in {sheet_name}: columns {start_col_idx}-{end_col_idx}, rows {start_row_idx}-{end_row_idx}")
            
            table_top_row, table_bottom_row = self._detect_table_boundaries(
                sheet, start_col_idx, end_col_idx, start_row_idx, end_row_idx
            )
            
            if not (table_top_row and table_bottom_row):
                logger.warning(f"No table boundaries found in page {page_num} of {sheet_name}")
                continue

            logger.info(f"Found table boundaries in page {page_num}: rows {table_top_row}-{table_bottom_row}")
            entries_in_page = 0

            for row in range(table_top_row + 1, table_bottom_row + 1):
                symbol = sheet.cell(row=row, column=start_col_idx).value
                course_name = sheet.cell(row=row, column=start_col_idx + 1).value
                teacher1 = sheet.cell(row=row, column=end_col_idx - 1).value
                teacher2 = sheet.cell(row=row, column=end_col_idx).value

                if symbol and course_name:
                    course_details.append([
                        symbol,
                        course_name, 
                        teacher1, 
                        teacher2, 
                        sheet_name
                    ])
                    entries_in_page += 1
            logger.info(f"--------------------------------------------------------------------------------")
            logger.info(f"Added {entries_in_page} course details from page {page_num} in {sheet_name}")
            logger.info(f"--------------------------------------------------------------------------------")
        logger.info(f"==============================================================================")
        logger.info(f"Total course details found in {sheet_name}: {len(course_details)}")
        logger.info(f"==============================================================================")
        return course_details

    def _parse_d_sheet(self, sheet, sheet_name: str, detail_origin: str) -> List[List]:
        """Parse a single d-sheet for courses"""
        courses = []
        pages = self._detect_pages(sheet)
        logger.info(f"Found {len(pages)} pages in course sheet: {sheet_name}")
        last_valid_month_year = None  # Store the last valid month/year across pages

        for page_num, (start_col_idx, end_col_idx, start_row_idx, end_row_idx) in enumerate(pages, 1):
            logger.info(f"Processing page {page_num} in {sheet_name}: columns {start_col_idx}-{end_col_idx}, rows {start_row_idx}-{end_row_idx}")
            
            table_top_row, table_bottom_row = self._detect_table_boundaries(
                sheet, start_col_idx, end_col_idx, start_row_idx, end_row_idx
            )
            
            if not (table_top_row and table_bottom_row):
                logger.warning(f"No table boundaries found in page {page_num} of {sheet_name}")
                continue

            logger.info(f"Found table boundaries in page {page_num}: rows {table_top_row}-{table_bottom_row}")
            entries_in_page = 0

            # Get month/year row before handling merged cells
            month_year_row = table_top_row
            week_row = table_top_row + 1
            date_row = table_top_row + 2

            # First, check and populate empty month cells at the start of the page
            if last_valid_month_year:
                for col in range(max(3, start_col_idx), end_col_idx + 1):
                    month_year = sheet.cell(row=month_year_row, column=col).value
                    if not month_year:  # Empty cell
                        # Only populate if we have a date in this column
                        date_str = sheet.cell(row=date_row, column=col).value
                        if date_str:
                            logger.info(f"Populating empty month cell in column {col} with {last_valid_month_year} on date {date_str}")
                            sheet.cell(row=month_year_row, column=col).value = last_valid_month_year
                            break
                    else:
                        # If we find a valid month/year, stop populating
                        year, month = self._parse_month_year(str(month_year))
                        if year and month:
                            break

            # Now handle merged cells
            self._handle_merged_cells(sheet, start_col_idx, end_col_idx, table_top_row, table_bottom_row)

            # Process the data and update last_valid_month_year
            for data_row in range(table_top_row + 3, table_bottom_row + 1):
                class_value = sheet.cell(row=data_row, column=1).value
                period_value = sheet.cell(row=data_row, column=2).value

                for col in range(max(3, start_col_idx), end_col_idx + 1):
                    month_year = sheet.cell(row=month_year_row, column=col).value
                    week = sheet.cell(row=week_row, column=col).value
                    date_str = sheet.cell(row=date_row, column=col).value
                    
                    
                    course_cell = sheet.cell(row=data_row, column=col)
                    course_value = course_cell.value
                    comment = course_cell.comment.text if course_cell.comment else ""

                    if all([month_year, date_str, course_value]):
                        year, month = self._parse_month_year(str(month_year))
                        last_valid_month_year = month_year
                        day = self._parse_date(str(date_str))
                        
                        if all([year, month, day]):
                            dt = datetime(year, month, day)
                            
                            # Extract event letter from course_value if present
                            course_symbol = str(course_value)
                            event = None
                            import re
                            match = re.match(r'^(\d{4})([A-Za-z]+)$', course_symbol)
                            if match:
                                course_symbol = match.group(1)  # The 4 digits
                                event = match.group(2).upper()       # The letters
                            
                            courses.append([
                                course_symbol, 
                                dt, 
                                week, 
                                class_value,
                                period_value, 
                                comment.rstrip(),
                                event,
                                sheet_name,
                                detail_origin
                            ])
                            entries_in_page += 1

            logger.info(f"Added {entries_in_page} courses from page {page_num} in {sheet_name}")

        logger.info(f"Total courses found in {sheet_name}: {len(courses)}")
        return courses

    def _detect_pages(self, sheet):
        """Detect page breaks in sheet"""
        vertical_breaks = [0] + [pb.id for pb in sheet.col_breaks.brk]
        horizontal_breaks = [0] + [pb.id for pb in sheet.row_breaks.brk]

        logger.info(f"Found {len(vertical_breaks)-1} vertical breaks and {len(horizontal_breaks)-1} horizontal breaks")

        if sheet.dimensions:
            last_cell = sheet.dimensions.split(":")[1]
            last_col = ''.join(re.findall(r'[A-Za-z]+', last_cell))
            last_row = int(''.join(re.findall(r'\d+', last_cell)))
        else:
            last_col = get_column_letter(sheet.max_column)
            last_row = sheet.max_row

        logger.info(f"Sheet dimensions: columns A-{last_col}, rows 1-{last_row}")
        pages = []
        for v in range(len(vertical_breaks)):
            end_col = get_column_letter(vertical_breaks[v+1] if v + 1 < len(vertical_breaks) else column_index_from_string(last_col))
            start_col = get_column_letter(vertical_breaks[v] + 1)
            start_col_idx = column_index_from_string(start_col)
            end_col_idx = column_index_from_string(end_col)

            for h in range(len(horizontal_breaks)):
                end_row_idx = horizontal_breaks[h+1] if h + 1 < len(horizontal_breaks) else last_row
                start_row_idx = horizontal_breaks[h] + 1
                pages.append((start_col_idx, end_col_idx, start_row_idx, end_row_idx))
        return pages

    def _detect_table_boundaries(self, sheet, start_col_idx, end_col_idx, start_row_idx, end_row_idx):
        """Detect table boundaries within a page"""
        table_top_row = None
        for row_idx in range(start_row_idx, end_row_idx + 1):
            cell = sheet.cell(row=row_idx, column=start_col_idx)
            if cell.border.top and cell.border.top.style is not None:
                table_top_row = row_idx
                break

        table_bottom_row = None
        for row_idx in range(end_row_idx, start_row_idx - 1, -1):
            cell = sheet.cell(row=row_idx, column=start_col_idx)
            if cell.border.bottom and cell.border.bottom.style is not None:
                table_bottom_row = row_idx
                break

        return table_top_row, table_bottom_row

    def _handle_merged_cells(self, sheet, start_col_idx, end_col_idx, table_top_row, table_bottom_row):
        """Handle merged cells in table"""
        for m_range in list(sheet.merged_cells.ranges):
            if (m_range.min_row >= table_top_row and m_range.max_row <= table_bottom_row and
                m_range.min_col >= start_col_idx and m_range.max_col <= end_col_idx):
                top_left_cell = sheet.cell(row=m_range.min_row, column=m_range.min_col)
                value = top_left_cell.value
                comment = top_left_cell.comment
                sheet.unmerge_cells(str(m_range))
                
                for r in range(m_range.min_row, m_range.max_row + 1):
                    for c in range(m_range.min_col, m_range.max_col + 1):
                        cell = sheet.cell(row=r, column=c)
                        cell.value = value
                        if comment:
                            cell.comment = type(comment)(text=comment.text, author=comment.author)

    def _parse_month_year(self, s):
        """Parse month and year from string"""
        parts = s.replace("THÁNG ", "").split("-")
        if len(parts) == 2:
            month_str = parts[0].strip()
            year_str = parts[1].strip()
            return int(year_str), int(month_str)
        return None, None

    def _parse_date(self, d_str):
        """Parse date from string"""
        try:
            return int(float(d_str))
        except:
            return None

    def store_data(self, courses: List[List], course_details: List[List]):
        """Store course and course detail data in the database"""
        # Clear existing data
        self.db_manager.execute_query('DELETE FROM courses')
        self.db_manager.execute_query('DELETE FROM course_details')
        
        # Store course details
        if course_details:
            # Check if the course_details list items have the correct number of elements
            # The SQL query expects 5 parameters: course_symbol, course_name, teacher_1, teacher_2, data_origin
            # But the processed_details might only have 4 elements if data_origin was removed
            
            # Check the first item to determine the format
            if course_details and len(course_details[0]) == 4:
                logger.info("Course details have 4 elements, adding data_origin")
                # Add data_origin as the 5th element if it's missing
                course_details_with_origin = []
                for detail in course_details:
                    # Add 'unknown' as the data_origin if it's missing
                    course_details_with_origin.append(detail + ['unknown'])
                
                self.db_manager.execute_many(
                    'INSERT INTO course_details (course_symbol, course_name, teacher_1, teacher_2, data_origin) VALUES (?, ?, ?, ?, ?)',
                    course_details_with_origin
                )
            else:
                # Use the details as is if they already have 5 elements
                self.db_manager.execute_many(
                    'INSERT INTO course_details (course_symbol, course_name, teacher_1, teacher_2, data_origin) VALUES (?, ?, ?, ?, ?)',
                    course_details
                )
            
            logger.info(f"Stored {len(course_details)} course details in database")

        # Store courses with event data
        if courses:
            self.db_manager.execute_many(
                'INSERT INTO courses (course_symbol, course_datetime, week, class, period, comment, event, data_origin, detail_origin) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)',
                courses
            )
            logger.info(f"Stored {len(courses)} courses in database")

    def export_to_csv(self, output_file: str = "course_schedule.csv"):
        """Export the database content to CSV"""
        query = '''
            SELECT 
                c.course_symbol,
                c.course_datetime,
                c.week,
                cd.teacher_1,
                cd.teacher_2,
                c.class,
                c.period,
                c.comment,
                c.event,
                c.data_origin,
                c.detail_origin,
                c.hall
            FROM courses c
            LEFT JOIN course_details cd 
            ON c.course_symbol = cd.course_symbol 
            AND c.detail_origin = cd.data_origin
            ORDER BY datetime(c.course_datetime)
        '''
        
        rows = self.db_manager.execute_query(query, fetch_all=True)

        with open(output_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow([
                "Course Symbol", "Date", "Week", "Teacher 1", "Teacher 2",
                "Class", "Period", "Comment", "Event", "Course Origin", "Detail Origin", "Hall"
            ])
            
            for row in rows:
                date_str = datetime.fromisoformat(row[1]).strftime("%d/%m/%Y") if row[1] else "N/A"
                writer.writerow([
                    row[0],
                    date_str,
                    row[2] or "N/A",
                    row[3] or "N/A",
                    row[4] or "N/A",
                    row[5] or "N/A",
                    row[6] or "N/A",
                    row[7] or "N/A",
                    row[8] or "N/A",
                    row[9],
                    row[10] or "No matching detail",
                    row[11] or "N/A"
                ])

        logger.info(f"Data exported to {output_file}")

    def parse_excel(self, filename: str) -> Tuple[List[List], List[List]]:
        """Parse Excel file with new matching strategy"""
        workbook = load_workbook(filename=filename)
        all_courses = []
        all_course_details = []

        # First pass: collect all Danh mục sheets and their details
        danh_muc_sheets = {}  # Map of sheet number to sheet name
        for sheet in workbook.worksheets:
            sheet_name = sheet.title
            if self._is_danh_muc_sheet(sheet_name):
                sheet_number = self._extract_sheet_number(sheet_name)
                if sheet_number != -1:
                    logger.info(f"Found Danh mục sheet: {sheet_name} (number: {sheet_number})")
                    danh_muc_sheets[sheet_number] = sheet_name
                    details = self._parse_course_details_sheet(sheet, sheet_name)
                    all_course_details.extend(details)

        # Second pass: process d-sheets only if matching Danh mục exists
        for sheet in workbook.worksheets:
            sheet_name = sheet.title
            if self._is_d_sheet(sheet_name):
                sheet_number = self._extract_sheet_number(sheet_name)
                if sheet_number != -1 and sheet_number in danh_muc_sheets:
                    logger.info(f"Processing d-sheet: {sheet_name} with matching Danh mục: {danh_muc_sheets[sheet_number]}")
                    courses = self._parse_d_sheet(sheet, sheet_name, danh_muc_sheets[sheet_number])
                    all_courses.extend(courses)
                else:
                    logger.warning(f"Skipping d-sheet {sheet_name} - no matching Danh mục sheet found")

        return all_courses, all_course_details

# Usage example:
if __name__ == "__main__":
    # Get the directory of the current script
    current_dir = os.path.dirname(os.path.abspath(__file__))
    # Construct the path to the data folder
    data_folder = os.path.join(current_dir, "data")
    # Construct the path to the input file
    input_file = os.path.join(data_folder, "test2.xlsx")

    cms = CourseManagementSystem("courses.db")
    courses, course_details = cms.parse_excel(input_file)
    cms.store_data(courses, course_details)
    cms.export_to_csv(output_file="courses.csv")